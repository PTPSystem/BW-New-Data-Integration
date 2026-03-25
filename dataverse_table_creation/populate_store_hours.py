#!/usr/bin/env python3
"""
Populate crf63_storeoperatinghour table in Dataverse
Imports Store Operating Hours from BI Dimensions Excel file

Table: BW Store Operating Hour
Schema Name: crf63_storeoperatinghour
Description: Master data for store operating hours by day of week

Columns:
  - crf63_storeoperatinghourid (GUID) - Primary key
  - crf63_storenumber (Whole number) - Store identifier
  - crf63_dayofweek (Whole number) - Day of week (1=Monday, 7=Sunday)
  - crf63_openingtimehhmm (String) - Opening time in HH:MM format (Primary Name)
  - crf63_closingtimehhmm (String) - Closing time in HH:MM format
  - crf63_openingtimedatetime (DateTime) - Opening time as datetime
  - crf63_closingtimedatetime (DateTime) - Closing time as datetime
  - crf63_businesskey (String) - Business Key: {StoreNumber}_{DayOfWeek} (for alternate key)

Data Source:
  - Excel file: BI Dimensions.xlsx
  - Sheet: Store hours
  - Location: OneDrive shared IT Project folder

Usage:
  python populate_store_hours.py [--excel-path PATH]
  
Options:
  --excel-path    Path to BI Dimensions.xlsx (optional, uses default shared OneDrive location)
  --dry-run       Show what would be imported without making changes
  --verbose       Show detailed progress information
"""

import json
import time
import argparse
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests
import msal
import openpyxl

# Add parent directory to path for imports
sys.path.append(str(Path(__file__).parent.parent))
from modules.utils.keyvault import get_dataverse_credentials

# ========================= CONFIG =========================
DEFAULT_EXCEL_PATH = "/Users/howardshen/Library/CloudStorage/OneDrive-SharedLibraries-globalpacmgt.com/IT Project - General/BI Import/BI Dimensions.xlsx"
SHEET_NAME = "Store hours"
TABLE_NAME = "crf63_storeoperatinghours"
TABLE_PLURAL = "crf63_storeoperatinghours"

# Expected Excel columns
EXCEL_COLUMNS = {
    'store_number': 0,      # Column A
    'day_of_week': 1,       # Column B
    'open_time': 2,         # Column C
    'close_time': 3         # Column D
}

# Days of week mapping
DAY_NAMES = {
    1: 'Monday',
    2: 'Tuesday', 
    3: 'Wednesday',
    4: 'Thursday',
    5: 'Friday',
    6: 'Saturday',
    7: 'Sunday'
}
# =========================================================


def get_auth_token():
    """Get authentication token for Dataverse using Azure Key Vault credentials"""
    try:
        creds = get_dataverse_credentials()
        app = msal.ConfidentialClientApplication(
            creds['client_id'],
            authority=f"https://login.microsoftonline.com/{creds['tenant_id']}",
            client_credential=creds['client_secret']
        )
        
        result = app.acquire_token_for_client(scopes=[f"{creds['environment_url']}/.default"])
        
        if "access_token" not in result:
            raise Exception(f"Failed to get token: {result.get('error_description', 'Unknown error')}")
        
        return result["access_token"], creds['environment_url']
    
    except Exception as e:
        print(f"ERROR: Authentication failed: {e}")
        sys.exit(1)


def convert_time_to_hhmm(time_value):
    """
    Convert various time formats to HH:MM string format
    
    Handles:
    - String formats: "1100", "11:00", "11 00"
    - Integer formats: 1100, 100
    - datetime.time objects
    - None/empty values
    
    Returns:
        str: Time in HH:MM format or None if invalid
    """
    if time_value is None or time_value == '':
        return None
    
    # If it's already a string, clean and format it
    if isinstance(time_value, str):
        # Remove any non-digits
        clean = ''.join(c for c in time_value if c.isdigit())
        if len(clean) == 0:
            return None
        elif len(clean) <= 2:
            # Format 1 -> 01:00, 12 -> 12:00
            return f"{int(clean):02d}:00"
        elif len(clean) == 3:
            # Format 100 -> 01:00, 930 -> 09:30
            return f"{int(clean[0]):02d}:{clean[1:3]}"
        elif len(clean) == 4:
            # Format 1100 -> 11:00
            return f"{clean[0:2]}:{clean[2:4]}"
        else:
            return None
    
    # If it's an integer
    if isinstance(time_value, int):
        time_str = str(time_value)
        if len(time_str) <= 2:
            return f"{time_value:02d}:00"
        elif len(time_str) == 3:
            return f"{int(time_str[0]):02d}:{time_str[1:3]}"
        elif len(time_str) == 4:
            return f"{time_str[0:2]}:{time_str[2:4]}"
    
    # If it's a datetime.time object
    if hasattr(time_value, 'hour') and hasattr(time_value, 'minute'):
        return f"{time_value.hour:02d}:{time_value.minute:02d}"
    
    return None


def convert_hhmm_to_datetime(time_str, base_date=None):
    """
    Convert HH:MM string to datetime object for Dataverse
    
    Args:
        time_str: Time in HH:MM format
        base_date: Base date to use (defaults to 2000-01-01)
    
    Returns:
        ISO 8601 formatted datetime string or None
    """
    if not time_str:
        return None
    
    if base_date is None:
        base_date = datetime(2000, 1, 1)
    
    try:
        parts = time_str.split(':')
        if len(parts) != 2:
            return None
        
        hour = int(parts[0])
        minute = int(parts[1])
        
        # Validate time components
        if hour < 0 or hour > 23 or minute < 0 or minute > 59:
            return None
        
        # Create datetime and convert to UTC
        dt = datetime(base_date.year, base_date.month, base_date.day, hour, minute, 0, tzinfo=timezone.utc)
        return dt.isoformat()
    
    except (ValueError, IndexError):
        return None


def load_excel_data(excel_path, verbose=False):
    """
    Load store hours data from BI Dimensions Excel file
    
    Args:
        excel_path: Path to Excel file
        verbose: Print detailed progress
    
    Returns:
        list: List of record dictionaries
    """
    print(f"\n{'='*60}")
    print(f"Loading Excel file")
    print(f"{'='*60}")
    print(f"Path: {excel_path}")
    
    if not Path(excel_path).exists():
        print(f"ERROR: Excel file not found: {excel_path}")
        sys.exit(1)
    
    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
        
        if SHEET_NAME not in wb.sheetnames:
            print(f"ERROR: Sheet '{SHEET_NAME}' not found in workbook")
            print(f"Available sheets: {', '.join(wb.sheetnames)}")
            sys.exit(1)
        
        ws = wb[SHEET_NAME]
        print(f"Sheet: {SHEET_NAME}")
        
    except Exception as e:
        print(f"ERROR: Failed to open Excel file: {e}")
        sys.exit(1)
    
    records = []
    headers = None
    skipped = 0
    
    for row_idx, row in enumerate(ws.iter_rows(values_only=True), 1):
        if row_idx == 1:
            # First row is headers
            headers = row
            if verbose:
                print(f"Headers: {headers}")
            continue
        
        # Skip empty rows
        if not row[EXCEL_COLUMNS['store_number']]:
            continue
        
        try:
            # Parse row data
            store_number = row[EXCEL_COLUMNS['store_number']]
            day_of_week = row[EXCEL_COLUMNS['day_of_week']]
            open_time_raw = row[EXCEL_COLUMNS['open_time']]
            close_time_raw = row[EXCEL_COLUMNS['close_time']]
            
            # Validate required fields
            if store_number is None or day_of_week is None:
                if verbose:
                    print(f"  Skipping row {row_idx}: Missing store number or day of week")
                skipped += 1
                continue
            
            # Convert to proper types
            store_number = int(store_number)
            day_of_week = int(day_of_week)
            
            # Validate day of week
            if day_of_week < 1 or day_of_week > 7:
                if verbose:
                    print(f"  Skipping row {row_idx}: Invalid day of week {day_of_week}")
                skipped += 1
                continue
            
            # Convert times to HH:MM format
            open_time = convert_time_to_hhmm(open_time_raw)
            close_time = convert_time_to_hhmm(close_time_raw)
            
            # Build Dataverse record
            record = {
                "crf63_storenumber": store_number,
                "crf63_dayofweek": day_of_week,
                "crf63_openingtimehhmm": open_time,
                "crf63_closingtimehhmm": close_time,
                "crf63_businesskey": f"{store_number}_{day_of_week}"
            }
            
            records.append(record)
            
            if verbose and len(records) % 10 == 0:
                print(f"  Loaded {len(records)} records...")
        
        except Exception as e:
            print(f"  ERROR on row {row_idx}: {e}")
            skipped += 1
            continue
    
    wb.close()
    
    print(f"\n{'='*60}")
    print(f"Excel Load Summary")
    print(f"{'='*60}")
    print(f"Total records loaded: {len(records)}")
    print(f"Rows skipped: {skipped}")
    
    # Show sample records
    if records:
        print(f"\nSample records (first 3):")
        for i, rec in enumerate(records[:3], 1):
            day_name = DAY_NAMES.get(rec['crf63_dayofweek'], 'Unknown')
            print(f"  {i}. Store {rec['crf63_storenumber']} - {day_name}: "
                  f"{rec.get('crf63_openingtimehhmm', 'N/A')} - {rec.get('crf63_closingtimehhmm', 'N/A')}")
    
    return records


def fetch_existing_records(session, api_url, verbose=False):
    """
    Fetch all existing store operating hour records from Dataverse
    
    Returns:
        dict: Mapping of "store_day" key to record ID
    """
    print(f"\n{'='*60}")
    print(f"Fetching existing records from Dataverse")
    print(f"{'='*60}")
    
    existing = {}
    url = f"{api_url}/{TABLE_PLURAL}?$select=crf63_storeoperatinghourid,crf63_storenumber,crf63_dayofweek"
    page = 0
    
    try:
        while url:
            page += 1
            if verbose:
                print(f"  Fetching page {page}...")
            
            resp = session.get(url, timeout=60)
            
            if resp.status_code != 200:
                print(f"ERROR: Failed to fetch records: HTTP {resp.status_code}")
                print(f"Response: {resp.text[:500]}")
                return {}
            
            data = resp.json()
            
            for record in data.get('value', []):
                store_num = record.get('crf63_storenumber')
                day = record.get('crf63_dayofweek')
                record_id = record.get('crf63_storeoperatinghourid')
                
                if store_num is not None and day is not None and record_id:
                    key = f"{store_num}_{day}"
                    existing[key] = record_id
            
            # Check for next page
            url = data.get('@odata.nextLink')
        
        print(f"Found {len(existing)} existing records in Dataverse")
        return existing
    
    except Exception as e:
        print(f"ERROR: Exception while fetching records: {e}")
        return {}


def upsert_record(session, api_url, record, existing_id=None):
    """
    Update existing record or create new one in Dataverse
    
    Args:
        session: requests.Session with auth headers
        api_url: Base API URL
        record: Record dictionary to upsert
        existing_id: GUID of existing record (if updating)
    
    Returns:
        tuple: (success: bool, status_code: int, message: str)
    """
    # Remove None values to avoid overwriting existing data with nulls
    clean_rec = {k: v for k, v in record.items() if v is not None}
    
    try:
        if existing_id:
            # UPDATE existing record
            url = f"{api_url}/{TABLE_PLURAL}({existing_id})"
            resp = session.patch(url, json=clean_rec, timeout=30)
            
            if resp.status_code in (200, 204):
                return True, resp.status_code, "updated"
            else:
                return False, resp.status_code, f"Update failed: {resp.text[:200]}"
        else:
            # CREATE new record
            url = f"{api_url}/{TABLE_PLURAL}"
            resp = session.post(url, json=clean_rec, timeout=30)
            
            if resp.status_code in (201, 204):
                return True, resp.status_code, "created"
            else:
                return False, resp.status_code, f"Create failed: {resp.text[:200]}"
    
    except Exception as e:
        return False, 0, f"Exception: {str(e)}"


def process_records(session, api_url, records, existing, dry_run=False, verbose=False):
    """
    Process all records with upsert logic
    
    Args:
        session: requests.Session with auth headers
        api_url: Base API URL
        records: List of records to process
        existing: Dict of existing record IDs
        dry_run: If True, only simulate without making changes
        verbose: Print detailed progress
    
    Returns:
        tuple: (success_count, update_count, create_count, error_count)
    """
    print(f"\n{'='*60}")
    print(f"Processing {len(records)} records" + (" [DRY RUN]" if dry_run else ""))
    print(f"{'='*60}")
    
    success_count = 0
    update_count = 0
    create_count = 0
    error_count = 0
    
    start_time = time.time()
    
    for idx, record in enumerate(records, 1):
        store_num = record['crf63_storenumber']
        day = record['crf63_dayofweek']
        day_name = DAY_NAMES.get(day, f"Day {day}")
        key = f"{store_num}_{day}"
        
        existing_id = existing.get(key)
        action = "UPDATE" if existing_id else "CREATE"
        
        if dry_run:
            # Simulate without making actual changes
            success_count += 1
            if existing_id:
                update_count += 1
            else:
                create_count += 1
            
            if verbose:
                print(f"  [{action}] Store {store_num} - {day_name}: "
                      f"{record.get('crf63_openingtimehhmm', 'N/A')} - "
                      f"{record.get('crf63_closingtimehhmm', 'N/A')}")
        else:
            # Actually perform the upsert
            success, status_code, message = upsert_record(session, api_url, record, existing_id)
            
            if success:
                success_count += 1
                if existing_id:
                    update_count += 1
                else:
                    create_count += 1
                
                if verbose or (idx % 25 == 0):
                    print(f"  Progress: {idx}/{len(records)} - {action} Store {store_num} - {day_name}")
            else:
                error_count += 1
                print(f"  ✗ ERROR: Store {store_num} - {day_name}: {message}")
        
        # Rate limiting (10 requests per second)
        if not dry_run and idx % 10 == 0:
            time.sleep(0.1)
    
    elapsed = time.time() - start_time
    
    print(f"\n{'='*60}")
    print(f"Processing Summary")
    print(f"{'='*60}")
    print(f"Total records processed: {len(records)}")
    print(f"Successful: {success_count}")
    print(f"  - Updated: {update_count}")
    print(f"  - Created: {create_count}")
    print(f"Failed: {error_count}")
    
    if not dry_run:
        print(f"Time elapsed: {elapsed:.1f} seconds")
        if elapsed > 0:
            print(f"Rate: {len(records) / elapsed:.1f} records/sec")
    
    return success_count, update_count, create_count, error_count


def main():
    """Main function to load store hours from Excel to Dataverse"""
    # Parse command line arguments
    parser = argparse.ArgumentParser(
        description='Populate Dataverse crf63_storeoperatinghour table from BI Dimensions Excel',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Use default Excel path
  python populate_store_hours.py
  
  # Use custom Excel path
  python populate_store_hours.py --excel-path /path/to/BI_Dimensions.xlsx
  
  # Dry run to see what would be imported
  python populate_store_hours.py --dry-run --verbose
  
  # Verbose output
  python populate_store_hours.py --verbose
        """
    )
    parser.add_argument('--excel-path', 
                       default=DEFAULT_EXCEL_PATH,
                       help='Path to BI Dimensions.xlsx file')
    parser.add_argument('--dry-run',
                       action='store_true',
                       help='Show what would be imported without making changes')
    parser.add_argument('--verbose', '-v',
                       action='store_true',
                       help='Show detailed progress information')
    
    args = parser.parse_args()
    
    print("="*60)
    print("Store Operating Hours Import to Dataverse")
    print("="*60)
    print(f"Table: {TABLE_NAME}")
    print(f"Mode: {'DRY RUN' if args.dry_run else 'LIVE'}")
    print(f"Verbose: {args.verbose}")
    
    overall_start = time.time()
    
    # Step 1: Load Excel data
    records = load_excel_data(args.excel_path, args.verbose)
    
    if not records:
        print("\nNo records to process!")
        return 1
    
    if args.dry_run:
        # In dry run mode, we don't need to authenticate or fetch existing records
        print(f"\n{'='*60}")
        print("DRY RUN - No changes will be made to Dataverse")
        print(f"{'='*60}")
        process_records(None, None, records, {}, dry_run=True, verbose=args.verbose)
    else:
        # Step 2: Authenticate with Dataverse
        print(f"\n{'='*60}")
        print("Authenticating with Dataverse")
        print(f"{'='*60}")
        token, environment_url = get_auth_token()
        api_url = f"{environment_url.rstrip('/')}/api/data/v9.2"
        print(f"Environment: {environment_url}")
        
        # Step 3: Setup session
        session = requests.Session()
        adapter = requests.adapters.HTTPAdapter(pool_connections=10, pool_maxsize=10)
        session.mount('https://', adapter)
        session.headers.update({
            "Authorization": f"Bearer {token}",
            "Content-Type": "application/json",
            "OData-MaxVersion": "4.0",
            "OData-Version": "4.0",
            "Prefer": "return=representation"
        })
        
        # Step 4: Fetch existing records
        existing = fetch_existing_records(session, api_url, args.verbose)
        
        # Step 5: Process records (upsert)
        success_count, update_count, create_count, error_count = process_records(
            session, api_url, records, existing, dry_run=False, verbose=args.verbose
        )
        
        # Final summary
        elapsed = time.time() - overall_start
        print(f"\n{'='*60}")
        print("Import Complete")
        print(f"{'='*60}")
        print(f"Total time: {elapsed:.1f} seconds")
        
        if error_count > 0:
            print(f"\n⚠️  {error_count} errors occurred during import")
            return 1
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
